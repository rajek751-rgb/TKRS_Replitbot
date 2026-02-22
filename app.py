import os
import re
import logging
from datetime import datetime
from flask import Flask, request
from telegram import (
    Update,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from config import TOKEN, GROUP_CHAT_ID, ADMIN_ID

logging.basicConfig(level=logging.INFO)

app_flask = Flask(__name__)
bot_app = Application.builder().token(TOKEN).build()

users = {}

TECH_LIST = [
    "ЦА","АЦН-10","АКН","АХО","ППУ","Цементосмеситель",
    "Автокран","Звено глушения","Звено СКБ","Тягач",
    "Седельный тягач","АЗА","Седельный тягач с КМУ",
    "Бортовой с КМУ","Топливозаправщик","Водовозка",
    "АРОК","Вахтовый автобус","УАЗ"
]

def valid_date(d):
    try:
        datetime.strptime(d, "%d.%m.%Y")
        return True
    except:
        return False

def valid_time(t):
    return re.match(r"^([01]\d|2[0-3]):([0-5]\d)$", t)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    users[update.effective_user.id] = {
        "step": "date",
        "operations": []
    }
    await update.message.reply_text("📅 Введите дату (ДД.ММ.ГГГГ)")

async def handle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    text = update.message.text
    user = users.get(uid)

    if not user:
        await update.message.reply_text("Введите /start")
        return

    # --- ДАТА ---
    if user["step"] == "date":
        if not valid_date(text):
            await update.message.reply_text("❌ Неверный формат даты")
            return
        user["date"] = text
        user["step"] = "shift"
        keyboard = [["I смена","II смена"],["Обе смены"]]
        await update.message.reply_text(
            "Выберите смену",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return

    # --- СМЕНА ---
    if user["step"] == "shift":
        user["shift"] = text
        user["step"] = "name"
        await update.message.reply_text("📝 Название операции", reply_markup=ReplyKeyboardRemove())
        return

    # --- НАЗВАНИЕ ---
    if user["step"] == "name":
        user["current"] = {"name": text}
        user["step"] = "start"
        await update.message.reply_text("⏰ Время начала (ЧЧ:ММ)")
        return

    # --- НАЧАЛО ---
    if user["step"] == "start":
        if not valid_time(text):
            await update.message.reply_text("❌ Неверный формат времени")
            return
        user["current"]["start"] = text
        user["step"] = "end"
        await update.message.reply_text("⏰ Время окончания (ЧЧ:ММ)")
        return

    # --- ОКОНЧАНИЕ ---
    if user["step"] == "end":
        if not valid_time(text):
            await update.message.reply_text("❌ Неверный формат времени")
            return
        if text <= user["current"]["start"]:
            await update.message.reply_text("❌ Окончание должно быть позже начала")
            return
        user["current"]["end"] = text
        user["step"] = "tech"
        keyboard = [TECH_LIST[i:i+3] for i in range(0, len(TECH_LIST), 3)]
        await update.message.reply_text(
            "🔧 Выберите технику",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return

    # --- ТЕХНИКА ---
    if user["step"] == "tech":
        user["current"]["tech"] = text
        user["step"] = "client"
        await update.message.reply_text("👤 Представитель заказчика")
        return

    # --- ПРЕДСТАВИТЕЛЬ ---
    if user["step"] == "client":
        user["current"]["client"] = text
        user["step"] = "materials"
        await update.message.reply_text("📦 Оборудование и материалы")
        return

    # --- МАТЕРИАЛЫ ---
    if user["step"] == "materials":
        user["current"]["materials"] = text
        user["operations"].append(user["current"])
        user["current"] = {}
        user["step"] = "action"

        keyboard = [["➕ Добавить ещё"],["✅ Завершить"]]
        await update.message.reply_text(
            "Операция добавлена",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return

    # --- ДЕЙСТВИЯ ---
    if user["step"] == "action":
        if text == "➕ Добавить ещё":
            user["step"] = "name"
            await update.message.reply_text("📝 Название операции")
            return

        if text == "✅ Завершить":
            filename = generate_excel(user)
            await update.message.reply_document(open(filename, "rb"))

            if GROUP_CHAT_ID:
                await context.bot.send_document(GROUP_CHAT_ID, open(filename, "rb"))

            if ADMIN_ID:
                await context.bot.send_document(ADMIN_ID, open(filename, "rb"))

            users[uid] = {}
            return

def generate_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Сетевой график"

    ws.merge_cells("A1:G1")
    ws["A1"] = "Сетевой график ТКРС"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["Дата:", data["date"]])
    ws.append(["Смена:", data["shift"]])
    ws.append([])

    ws.append([
        "№",
        "Наименование",
        "Начало",
        "Окончание",
        "Техника",
        "Представитель",
        "Оборудование"
    ])

    for i, op in enumerate(data["operations"], 1):
        ws.append([
            i,
            op["name"],
            op["start"],
            op["end"],
            op["tech"],
            op["client"],
            op["materials"]
        ])

    filename = f"TKRS_{data['date']}.xlsx"
    wb.save(filename)
    return filename

bot_app.add_handler(CommandHandler("start", start))
bot_app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle))

@app_flask.route(f"/{TOKEN}", methods=["POST"])
async def webhook():
    update = Update.de_json(request.get_json(force=True), bot_app.bot)
    await bot_app.process_update(update)
    return "ok"

@app_flask.route("/")
def home():
    return "Enterprise TKRS Bot Running"

if __name__ == "__main__":
    bot_app.run_polling()